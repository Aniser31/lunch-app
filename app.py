from flask import Flask, render_template, request, redirect, url_for, session, send_file, flash
from datetime import datetime
import os
import psycopg2
import psycopg2.extras
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "supersecretkey")  # Required for sessions

# ===============================
# Admin credentials
# ===============================
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "1234")

# ===============================
# Persistent storage (PostgreSQL)
# ===============================
DATABASE_URL = os.getenv("DATABASE_URL")

# Railway sometimes gives postgres://, psycopg2 needs postgresql://
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

def get_conn():
    db_url = os.getenv("DATABASE_URL")
    if not db_url:
        raise RuntimeError("DATABASE_URL is not set. Did you add it in Railway?")
    if db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql://", 1)
    return psycopg2.connect(db_url, sslmode="require")

def init_db():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS orders (
            id SERIAL PRIMARY KEY,
            doc TEXT NOT NULL,
            leader TEXT NOT NULL,
            member TEXT NOT NULL,
            vendor TEXT NOT NULL,
            menu TEXT NOT NULL,
            date DATE NOT NULL,
            UNIQUE(member, date)
        )
    """)
    conn.commit()
    cur.close()
    conn.close()

def add_or_update_order(order):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO orders (doc, leader, member, vendor, menu, date)
        VALUES (%s, %s, %s, %s, %s, %s)
        ON CONFLICT (member, date) DO UPDATE SET
            doc = EXCLUDED.doc,
            leader = EXCLUDED.leader,
            vendor = EXCLUDED.vendor,
            menu = EXCLUDED.menu
    """, (order["doc"], order["leader"], order["member"], order["vendor"], order["menu"], order["date"]))
    conn.commit()
    cur.close()
    conn.close()

def get_orders(start_date=None, end_date=None):
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    query = "SELECT id, doc, leader, member, vendor, menu, date FROM orders"
    params = []
    clauses = []
    if start_date:
        clauses.append("date >= %s")
        params.append(start_date)
    if end_date:
        clauses.append("date <= %s")
        params.append(end_date)
    if clauses:
        query += " WHERE " + " AND ".join(clauses)
    query += " ORDER BY date ASC, id ASC"
    cur.execute(query, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [dict(r) for r in rows]

def delete_order(order_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM orders WHERE id = %s", (order_id,))
    conn.commit()
    cur.close()
    conn.close()

def clear_orders():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM orders")
    conn.commit()
    cur.close()
    conn.close()

# ===============================
# Data
# ===============================
team_data = {
    "Suyogya": {
        "Suyogya": ["Suyogya", "Ranju Maharjan", "Ishwor Acharya", "Ujwal Shrestha", "Amrita Kumari Shah", "Rashmi Maharjan", "Dilip Khanal"],
        "Farukh": ["Farukh", "Shresna Maharjan", "Jyoti Sardar", "Barsha Raskoti", "Subash Bishunke", "Ronish Makaju", "Inghang Limbu"],
        "Santosh Sah(M)": ["Santosh Sah", "Harish Bhatta", "Sumitra Gyapak", "Utsav Goja", "Gaurav Dhakal", "Jasmine Dhoju", "Prakash Khati"]
    },
    "Darshan": {
        "Darshan": ["Darshan", "Bijay Yonjan", "Saroj Suwal", "Astup Dhju", "Nisha Koju", "Rabin Koju", "Seesam Maharjan"],
        "Ishwor": ["Ishwor", "Purna Bahadur Shahi", "Kabi raj bhatta", "Nawanit Madhikarmy", "Anurup Chaudhary", "Manisha Bhujel", "Shishir Kandel"],
        "Nabin": ["Nabin", "Puja Chaulagain", "Rupesh Gurung", "Ashish Kumar Sah", "Bina Achhami", "Suraj Devkota", "Alish Poudel"]
    },
    "Bishal": {
        "Bishal": ["Bishal", "Geeta Prajapati", "Kiran chaudhary", "Subash Thapa", "Rajan Limbu", "Ram Tamang", "Rohan Rajchal"],
        "Jayanti": ["Jayanti", "Saloni Shrestha", "Anita Ramauli", "Nirnaya Pandey", "Samikshya Adhikari", "Sajana Khyaju", "Sachita Chaudhary"],
        "Nirjal": ["Nirjal", "Pujan Shrestha", "Bibesh Rai", "Srijana Magar", "Rupa Thokar", "Sharmila Dhami", "Saricha Gautam"]
    },
    "Anjana": {
        "Anjana": ["Anjana", "Sumesh Khoju", "Manish Chaudhary", "Suchana GC", "Sudeep", "Karan Achhami"],
        "Rajesh": ["Rajesh", "Mina Bogati", "Birat Laudari", "Sagar Regmi", "Ashok Makaju", "Rojan Shrestha"],
        "Jeevan": ["Jeevan", "Gaurav Ale Magar", "Bikal Jadali"]
    },
    "Puskar": {
        "Puskar": ["Puskar", "Anubhav Pancha", "Sunita" , "Bhakta Achhami", "Binod Dhakal", "Sanish Shrestha"],
        "Biwas": ["Biwas", "Roshan Pun", "Sadhana Kumari Ray", "Sadipa Dhakal", "Sunita Kumal", "Sushma Achhami", "Binita Gora"],
        "Bibek": ["Bibek", "Dhan Bahadur BK", "Sikha", "Unika Maharjan", "Abhishek Karki", "Raj Bishunke"]
    },
    "Rukesh": {
        "Rukesh": ["Rukesh", "Bibek Budha", "Bibita Bati", "Amrit Dhakal", "Sanjok Khadka", "Kriti"],
        "Madan Shrestha": ["Madan Shrestha", "Bishal Achhami", "Anil Lakhaju", "Rocky Suwal", "Rahul Garu", "Raskin Baiju", "Rohan Bahala", "Sabina Dhamala"]
    },
    "Others": {
        "Others": ["Dishoj Sir", "Nilesh Sir", "Saugat Sir", "Manita Budhathoki", "Deebin Shrestha", "Sashank Sir", "Aditya Chaudhary", 
        "Rejina", "Enjeela Chaudhary", "Raunak Subedi", "Niru Dhaubanjar", "Sujal Shrestha", "Ashant Chaudhary", "Arun Mahara", "Simon Pulami", "Labin Sir", "Bibek Tamang"]
    },
    "Pramod Niraula":{
        "Pramod": ["Aashish Lama", "Puja Kandel", "Swwosti Adhikari"],
        "Puja": ["Puja Yadav", "Ashwinee Poudel", "Swornima Chaudhary", "Gokul Budha"],
        "Ashish":["Ashish Chantyal", "AR Ramesh"],
        "Sarina": ["Sarina Manandhar", "Debit Chaudhary"]
    }
}

vendor_menus = {
    "Vendor 1": [
        "Momo Veg", "Momo Chi", "Momo Buff",
        "Chowmein Veg", "Chowmein Chi", "Chowmein Buff",
        "Fried Rice Veg", "Fried Rice Chi", "Fried Rice Buff",
        "Burger Veg", "Burger Chi",
        "Sandwich Veg", "Sandwich Chi",
        "Rice w butter chicken", "Rice w paneer tofu",
        "Curry Veg", "Curry Chi"
    ],
    "Vendor 2": [
        "Non veg Khana set", "Veg Khana set"
    ]
}

# ===============================
# Template helpers
# ===============================
def all_menu_items():
    items = set()
    for arr in vendor_menus.values():
        items.update(arr)
    return sorted(items)

# ===============================
# Excel exports
# ===============================
ORDER_PRICE = 85

def generate_orders_excel(orders_to_export, team_data):
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    if not orders_to_export:
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio

    df = pd.DataFrame(orders_to_export)
    df["date_dt"] = pd.to_datetime(df["date"], format="%Y-%m-%d")
    df["formatted_date"] = df["date_dt"].dt.strftime("%B %d")

    summary = df.groupby(["formatted_date", "doc"]).size().unstack(fill_value=0)
    date_order = sorted(df["formatted_date"].unique(), key=lambda d: datetime.strptime(d, "%B %d"))
    summary = summary.reindex(date_order)

    totals = summary.sum(axis=0)
    summary.loc["Total Orders"] = totals
    summary.loc["Total Price (Rs)"] = totals * ORDER_PRICE

    for r in dataframe_to_rows(summary.reset_index(), index=False, header=True):
        ws_summary.append(r)

    # Per-DOC member pages
    for doc in team_data.keys():
        ws = wb.create_sheet(title=doc[:31])  # Excel sheet name limit 31
        ws.append(["Member", "Order Count", "Total Price (Rs)"])
        doc_orders = df[df["doc"] == doc]
        if doc_orders.empty:
            continue
        member_counts = doc_orders.groupby("member").size()
        for member, count in member_counts.items():
            ws.append([member, int(count), int(count) * ORDER_PRICE])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

def generate_food_orders_excel(orders_to_export):
    """
    Creates a pivot-style report:
      - "Food Orders" sheet: DOCs vs Food Items
      - "Total" sheet: Food Item totals across all DOCs
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Food Orders"

    if not orders_to_export:
        # empty workbook
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio

    df = pd.DataFrame(orders_to_export)

    # Normalize DOC name: show 'Admin' for IT/Admin/Managers
    def map_doc(d):
        return "Admin" if d == "IT/Admin/Managers" else d

    df["DOC_LABEL"] = df["doc"].apply(map_doc)

    # Columns list includes all known menu items, plus any ad-hoc items found
    menu_cols = set(all_menu_items())
    menu_cols.update(df["menu"].unique().tolist())
    menu_cols = sorted(menu_cols)

    pivot = df.pivot_table(
        index="DOC_LABEL",
        columns="menu",
        values="member",
        aggfunc="count",
        fill_value=0
    )

    # Ensure all expected menu columns exist
    for col in menu_cols:
        if col not in pivot.columns:
            pivot[col] = 0
    pivot = pivot[menu_cols]  # order columns

    # Add row totals
    pivot["Total per DOC"] = pivot.sum(axis=1)

    # Add grand total row
    grand_totals = pivot.sum(axis=0)
    pivot.loc["Grand Total"] = grand_totals

    # Write to Food Orders sheet
    ws.append(["DOC"] + list(pivot.columns))
    for idx, row in pivot.iterrows():
        ws.append([idx] + list(row.values))

    # ---------- Add TOTAL sheet ----------
    ws_total = wb.create_sheet(title="Total")
    ws_total.append(["Menu Item", "Total Orders"])

    total_counts = df["menu"].value_counts().reindex(menu_cols, fill_value=0)
    for item, count in total_counts.items():
        ws_total.append([item, int(count)])

    ws_total.append(["Grand Total", int(total_counts.sum())])

    # Save workbook
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ===============================
# Routes
# ===============================
@app.route("/", methods=["GET", "POST"])
def order():
    selected_doc = request.form.get("doc", "") or request.args.get("doc", "")
    selected_leader = request.form.get("leader", "") or request.args.get("leader", "")
    selected_member = request.form.get("member", "") or request.args.get("member", "")
    selected_vendor = request.form.get("vendor", "") or request.args.get("vendor", "")
    selected_date = request.form.get("date", "") or request.args.get("date", "")

    if request.method == "POST" and all(k in request.form for k in ("doc", "leader", "member", "vendor", "menu", "date")):
        add_or_update_order({
            "doc": request.form["doc"],
            "leader": request.form["leader"],
            "member": request.form["member"],
            "vendor": request.form["vendor"],
            "menu": request.form["menu"],
            "date": request.form["date"]
        })
        flash("Order placed/updated successfully.", "success")
        return redirect(url_for('order'))

    # For the public page we still show existing orders (no admin controls)
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    orders = get_orders(start_date if start_date else None, end_date if end_date else None)

    return render_template(
        "index.html",
        team_data=team_data,
        vendor_menus=vendor_menus,
        selected_doc=selected_doc,
        selected_leader=selected_leader,
        selected_member=selected_member,
        selected_vendor=selected_vendor,
        selected_date=selected_date,
        start_date=start_date,
        end_date=end_date,
        orders=orders
    )

# ---------- Admin ----------
@app.route("/admin", methods=["GET"])
def admin():
    if not session.get("admin"):
        return redirect(url_for("admin_login"))
    # Defaults for dashboard filter
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    orders = get_orders(start_date if start_date else None, end_date if end_date else None)
    return render_template("admin_dashboard.html", orders=orders, start_date=start_date, end_date=end_date)

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session["admin"] = True
            flash("Logged in successfully.", "success")
            return redirect(url_for("admin"))
        else:
            flash("Invalid credentials.", "danger")
    return render_template("admin_login.html")

@app.route("/admin/logout", methods=["POST"])
def admin_logout():
    session.pop("admin", None)
    flash("Logged out.", "info")
    return redirect(url_for("order"))

@app.route("/delete/<int:order_id>", methods=["POST"])
def delete(order_id):
    if session.get("admin"):
        delete_order(order_id)
        flash("Order deleted.", "info")
    return redirect(request.referrer or url_for("admin"))

@app.route("/clear", methods=["POST"])
def clear_all():
    if session.get("admin"):
        clear_orders()
        flash("All orders cleared.", "warning")
    return redirect(request.referrer or url_for("admin"))

# ---------- Exports ----------
@app.route("/export-excel")
def export_excel():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))

    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    orders_to_export = get_orders(start_date if start_date else None, end_date if end_date else None)
    excel_buffer = generate_orders_excel(orders_to_export, team_data)

    return send_file(
        excel_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='lunch_orders_summary.xlsx'
    )

@app.route("/export-food-excel")
def export_food_excel():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))

    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    orders_to_export = get_orders(start_date if start_date else None, end_date if end_date else None)
    excel_buffer = generate_food_orders_excel(orders_to_export)

    return send_file(
        excel_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='lunch_food_orders.xlsx'
    )

# ===============================
# Start
# ===============================
# Always run init_db on startup (important for Railway/Gunicorn)
init_db()

if __name__ == "__main__":
    # Use PORT from Railway; enable debug locally
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)
